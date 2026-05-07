#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict

import pandas as pd
import requests

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "output"
CACHE = ROOT / ".cache"
OUTPUT.mkdir(exist_ok=True)
CACHE.mkdir(exist_ok=True)
LOG_PATH = OUTPUT / "run_log.txt"


def log(msg: str) -> None:
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line, flush=True)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_config(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def valid_csv(path: Path, series_id: str) -> bool:
    if not path.exists() or path.stat().st_size < 20:
        return False
    head = path.read_text(encoding="utf-8-sig", errors="replace")[:1000]
    return (("DATE" in head or "date" in head or "observation_date" in head)
            and (series_id in head or "VALUE" in head or "value" in head))


def download_series(series_id: str, url: str, out_path: Path, timeout: int, retries: int, sleep: int):
    headers = {
        "User-Agent": "Mozilla/5.0 macro-tracker-github-actions",
        "Accept": "text/csv,text/plain,*/*",
        "Connection": "close",
    }
    last_msg = ""
    for i in range(1, retries + 1):
        try:
            log(f"Downloading {series_id}: attempt {i}/{retries}")
            with requests.get(url, headers=headers, timeout=timeout, stream=True) as r:
                r.raise_for_status()
                tmp = out_path.with_suffix(".tmp")
                with tmp.open("wb") as f:
                    for chunk in r.iter_content(chunk_size=65536):
                        if chunk:
                            f.write(chunk)
                tmp.replace(out_path)
            if valid_csv(out_path, series_id):
                return True, "Downloaded successfully"
            last_msg = "Downloaded file did not match expected CSV format"
        except Exception as e:
            last_msg = f"{type(e).__name__}: {e}"
            log(f"{series_id}: download failed: {last_msg}")
            if i < retries:
                time.sleep(sleep)
    return False, last_msg


def get_series_csv(series: dict, settings: dict):
    sid = series["id"]
    url = series["url"]
    cache_path = CACHE / f"{sid}.csv"
    status = {
        "Series": sid, "Status": "", "SourceUsed": "", "Message": "",
        "Url": url, "Timestamp": datetime.now().isoformat(timespec="seconds")
    }
    ok, msg = download_series(
        sid, url, cache_path,
        int(settings.get("request_timeout_seconds", 30)),
        int(settings.get("max_retries", 4)),
        int(settings.get("retry_sleep_seconds", 3)),
    )
    if ok:
        status.update(Status="OK", SourceUsed="online_requests", Message=msg)
        return cache_path, status
    if valid_csv(cache_path, sid):
        status.update(Status="STALE_OK", SourceUsed="cache", Message=f"Online failed; using existing cache: {msg}")
        return cache_path, status
    status.update(Status="FAILED", SourceUsed="none", Message=f"Online failed and no cache exists: {msg}")
    return None, status


def read_series(path: Optional[Path], sid: str) -> pd.DataFrame:
    if path is None or not path.exists():
        return pd.DataFrame(columns=["date", sid])
    try:
        df = pd.read_csv(path, comment="#")
    except Exception as e:
        log(f"{sid}: CSV read failed: {e}")
        return pd.DataFrame(columns=["date", sid])
    date_col = next((c for c in ["DATE", "date", "observation_date"] if c in df.columns), None)
    val_col = sid if sid in df.columns else next((c for c in ["VALUE", "value"] if c in df.columns), None)
    if date_col is None or val_col is None:
        log(f"{sid}: could not find date/value columns")
        return pd.DataFrame(columns=["date", sid])
    out = df[[date_col, val_col]].copy()
    out.columns = ["date", sid]
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out[sid] = pd.to_numeric(out[sid], errors="coerce")
    return out.dropna(subset=["date", sid]).sort_values("date")


def latest_at_or_before(df: pd.DataFrame, sid: str, date: pd.Timestamp):
    if df.empty:
        return None
    sub = df[df["date"] <= date]
    return None if sub.empty else sub.iloc[-1]


def yoy(df: pd.DataFrame, sid: str, obs) -> Optional[float]:
    if obs is None or df.empty:
        return None
    target = obs["date"] - pd.DateOffset(years=1) + pd.Timedelta(days=15)
    prior_df = df[df["date"] <= target]
    if prior_df.empty:
        return None
    prior = prior_df.iloc[-1][sid]
    if pd.isna(prior) or prior == 0:
        return None
    return (float(obs[sid]) / float(prior) - 1.0) * 100.0


def qoq_annualized(df: pd.DataFrame, sid: str, obs) -> Optional[float]:
    if obs is None or df.empty:
        return None
    prior_df = df[df["date"] < obs["date"]]
    if prior_df.empty:
        return None
    prior = prior_df.iloc[-1][sid]
    if pd.isna(prior) or prior == 0:
        return None
    return ((float(obs[sid]) / float(prior)) ** 4 - 1.0) * 100.0


def change_3m(df: pd.DataFrame, sid: str, obs) -> Optional[float]:
    if obs is None or df.empty:
        return None
    target = obs["date"] - pd.DateOffset(months=3) + pd.Timedelta(days=15)
    prior_df = df[df["date"] <= target]
    if prior_df.empty:
        return None
    return float(obs[sid]) - float(prior_df.iloc[-1][sid])


def n(x):
    if x is None or pd.isna(x):
        return None
    return round(float(x), 2)


def determine_phase(score: int, activity_yoy, activity_chg) -> str:
    if score <= 1:
        return "Pessimism"
    if score == 2:
        return "Disbelief / Early Recovery"
    if score == 3:
        return "Skepticism"
    if score >= 4:
        if activity_yoy is not None and activity_yoy > 2 and activity_chg is not None and activity_chg < 0:
            return "Late Optimism / Chasing Risk"
        return "Optimism"
    return "Unknown"


def build_tracker(data: Dict[str, pd.DataFrame], months: int) -> pd.DataFrame:
    start = pd.Timestamp.today().normalize().replace(day=1) - pd.DateOffset(months=months)
    month_starts = pd.date_range(start=start, periods=months + 1, freq="MS")
    rows = []
    for m in month_starts:
        month_end = m + pd.offsets.MonthEnd(0)
        loan_obs = latest_at_or_before(data["BUSLOANS"], "BUSLOANS", month_end)
        gdp_obs = latest_at_or_before(data["GDPC1"], "GDPC1", month_end)
        act_obs = latest_at_or_before(data["IPMAN"], "IPMAN", month_end)

        loan_yoy = yoy(data["BUSLOANS"], "BUSLOANS", loan_obs)
        gdp_yoy = yoy(data["GDPC1"], "GDPC1", gdp_obs)
        gdp_qoq = qoq_annualized(data["GDPC1"], "GDPC1", gdp_obs)
        activity_yoy = yoy(data["IPMAN"], "IPMAN", act_obs)
        activity_chg_3m = change_3m(data["IPMAN"], "IPMAN", act_obs)

        score = 0
        if loan_yoy is not None and loan_yoy > 0:
            score += 1
        if loan_yoy is not None and gdp_yoy is not None and loan_yoy > gdp_yoy:
            score += 1
        if activity_yoy is not None and activity_yoy > 0:
            score += 1
        if activity_chg_3m is not None and activity_chg_3m > 0:
            score += 1
        if gdp_yoy is not None and gdp_yoy > 0:
            score += 1

        phase = determine_phase(score, activity_yoy, activity_chg_3m)

        rows.append({
            "Month": m.strftime("%Y-%m"),
            "Loans_Obs_Date": "" if loan_obs is None else loan_obs["date"].strftime("%Y-%m-%d"),
            "Loans_Value": None if loan_obs is None else n(loan_obs["BUSLOANS"]),
            "Loans_YoY": n(loan_yoy),
            "GDP_Obs_Date": "" if gdp_obs is None else gdp_obs["date"].strftime("%Y-%m-%d"),
            "GDP_Value": None if gdp_obs is None else n(gdp_obs["GDPC1"]),
            "GDP_YoY": n(gdp_yoy),
            "GDP_QoQ_Annualized": n(gdp_qoq),
            "Activity_Obs_Date": "" if act_obs is None else act_obs["date"].strftime("%Y-%m-%d"),
            "Activity_Proxy": "IPMAN",
            "Activity_Value": None if act_obs is None else n(act_obs["IPMAN"]),
            "Activity_YoY": n(activity_yoy),
            "Activity_3M_Change": n(activity_chg_3m),
            "Score": score,
            "Sentiment_Phase": phase,
        })
    return pd.DataFrame(rows)


def write_dashboard(tracker: pd.DataFrame, status_df: pd.DataFrame) -> None:
    latest = tracker.iloc[-1].to_dict() if not tracker.empty else {}
    recent_html = tracker.tail(30).to_html(index=False, border=0)
    status_html = status_df.to_html(index=False, border=0)
    html = f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>Macro Tracker Dashboard</title>
<style>
body {{ font-family: Segoe UI, Arial, sans-serif; margin: 24px; color: #222; }}
.card {{ padding: 14px; border: 1px solid #ddd; border-radius: 10px; margin: 12px 0; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 10px; font-size: 13px; }}
th, td {{ border: 1px solid #ddd; padding: 7px; text-align: left; }}
th {{ background: #f2f2f2; }}
.note {{ color: #555; font-size: 13px; }}
.good {{ color:#047857; font-weight:600; }}
.bad {{ color:#B91C1C; font-weight:600; }}
</style>
</head>
<body>
<h1>Macro Tracker Dashboard</h1>
<p class="note">Generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
<div class="card">
<h2>Latest signal</h2>
<p><b>Phase:</b> {latest.get("Sentiment_Phase","")} / <b>Score:</b> {latest.get("Score","")}</p>
<p><b>Loans YoY:</b> {latest.get("Loans_YoY","")} (obs {latest.get("Loans_Obs_Date","")}) |
<b>Activity:</b> {latest.get("Activity_Value","")} / YoY {latest.get("Activity_YoY","")} (obs {latest.get("Activity_Obs_Date","")}) |
<b>GDP YoY:</b> {latest.get("GDP_YoY","")} (obs {latest.get("GDP_Obs_Date","")})</p>
<p class="note">Activity proxy uses FRED IPMAN, not ISM/S&P PMI. IPMAN is hard manufacturing production data, not a diffusion PMI.</p>
</div>
<div class="card"><h2>Download status</h2>{status_html}</div>
<div class="card"><h2>Recent tracker</h2>{recent_html}</div>
</body>
</html>"""
    (OUTPUT / "dashboard.html").write_text(html, encoding="utf-8")
    # For GitHub Pages root URL
    (OUTPUT / "index.html").write_text(html, encoding="utf-8")


def write_excel(tracker: pd.DataFrame, summary: pd.DataFrame, status_df: pd.DataFrame) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.chart import LineChart, Reference
    except Exception as e:
        log(f"openpyxl not available; skipping xlsx creation: {e}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"
    for r in dataframe_to_rows(tracker, index=False, header=True):
        ws.append(r)

    ws2 = wb.create_sheet("Summary")
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws2.append(r)

    ws3 = wb.create_sheet("Download_Status")
    for r in dataframe_to_rows(status_df, index=False, header=True):
        ws3.append(r)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in [ws, ws2, ws3]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 32)
        sheet.freeze_panes = "A2"

    if ws.max_row > 5:
        chart = LineChart()
        chart.title = "Loans YoY / Activity YoY / GDP YoY"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Month"
        for col in [4, 7, 12]:
            data = Reference(ws, min_col=col, min_row=1, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 20
        ws.add_chart(chart, "Q2")

    out = ROOT / "macro_tracker.xlsx"
    wb.save(out)
    log(f"Created Excel workbook: {out}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=str(ROOT / "config.json"))
    args = parser.parse_args()

    LOG_PATH.write_text("", encoding="utf-8")
    config = load_config(Path(args.config))
    settings = config.get("settings", {})

    statuses = []
    data = {}
    for series in config["series"]:
        path, status = get_series_csv(series, settings)
        statuses.append(status)
        data[series["id"]] = read_series(path, series["id"])

    tracker = build_tracker(data, int(settings.get("history_months", 84)))
    tracker.to_csv(OUTPUT / "tracker.csv", index=False, encoding="utf-8-sig")

    status_df = pd.DataFrame(statuses)
    status_df.to_csv(OUTPUT / "download_status.csv", index=False, encoding="utf-8-sig")

    latest = tracker.iloc[-1] if not tracker.empty else {}
    summary = pd.DataFrame([
        {"Metric": "Sentiment Phase", "Value": latest.get("Sentiment_Phase", ""), "Note": "Rule-based phase from credit/GDP/activity proxy"},
        {"Metric": "Score", "Value": latest.get("Score", ""), "Note": "0 to 5"},
        {"Metric": "Loans YoY", "Value": latest.get("Loans_YoY", ""), "Note": f"Observed: {latest.get('Loans_Obs_Date', '')}"},
        {"Metric": "Activity Proxy", "Value": "IPMAN", "Note": f"Value: {latest.get('Activity_Value', '')}; YoY: {latest.get('Activity_YoY', '')}; Observed: {latest.get('Activity_Obs_Date', '')}"},
        {"Metric": "GDP YoY", "Value": latest.get("GDP_YoY", ""), "Note": f"Observed: {latest.get('GDP_Obs_Date', '')}"},
    ])
    summary.to_csv(OUTPUT / "summary.csv", index=False, encoding="utf-8-sig")

    write_dashboard(tracker, status_df)
    write_excel(tracker, summary, status_df)

    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "outputs": [
            "index.html",
            "dashboard.html",
            "tracker.csv",
            "summary.csv",
            "download_status.csv",
            "run_log.txt"
        ],
        "note": "output/index.html is the same dashboard for GitHub Pages root deployment."
    }
    (OUTPUT / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    log("Done. Pages site root will use output/index.html if deployed with GitHub Pages Actions.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
